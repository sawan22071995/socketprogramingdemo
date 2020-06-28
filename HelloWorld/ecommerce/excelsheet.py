#to import excel spreadsheet working module..
import openpyxl as xl
from openpyxl.chart import BarChart,Reference

#it is used to load workbook/spreadsheet file with name
wb=xl.load_workbook('transaction2.xlsx')

#it is used to load sheet from the workbook
sheet=wb['Sheet1']

#it is used to access the cell[column_name,row_name] from sheet
cell=sheet['a1']

#it is used to access the cell by get method(coloumn_no,row_no)
cell=sheet.cell(1,1)

#it is used to print the cell value
print(cell.value)

#it is used to find the maximum no of used/data value rows in sheet
print(sheet.max_row)

#it is used to find the maximum no of used/data value colomn in sheet
print(sheet.max_column)

#it is used to navigate between rows start from second row to last data value row in sheet
for row in range(2,sheet.max_row+1):

    #it is used to access 3rd column of each row
    cell=sheet.cell(row,3)

    #it is used to store new value in object
    price_new=cell.value*0.9

    #it is used to add new column access in sheet
    price_new_cell=sheet.cell(row,4)

    #it is used to add the value in that 4th(new column)
    price_new_cell.value=price_new

#add the value to particular cell with some value
sheet.cell(1,4).value='correct_price'

#add the blank value to particular cell
sheet.cell(1,5).value=''

#it is used to select the value from the sheet for chart representation and store in variable 'value'
value=Reference(sheet,
                min_row=2,
                max_row=sheet.max_row,
                min_col=4,
                max_col=4)

#is is used to create the instance of bar chart
chart=BarChart()

#its used to pass the value to chart for craetion
chart.add_data(value)

#it is used to insert chart at particular location into the sheet
sheet.add_chart(chart,'e2')
#it is used to save the sheet with same name or new sheet with new name
wb.save('transaction2.xlsx')

